from openpyxl import load_workbook

def write_api_answers_to_excel(
    header: str,
    values: list[str|int],
    excel_path: str,
):
    # 既存のExcelファイルを読み込む
    workbook = load_workbook(excel_path)
    sheet = workbook.active  # 操作対象のシートを指定 (デフォルトでアクティブシート)

    # 次の空列を探す
    next_column = sheet.max_column + 1  # 現在の最大列の次の列を取得

    # 行に沿って書き込み
    sheet.cell(row=1, column=next_column, value=header)
    for row_index, value in enumerate(values, start=2):  # 行番号は2から始まる
        sheet.cell(row=row_index, column=next_column, value=value)

    # ファイルを保存
    workbook.save(excel_path)
    print(f"データを列 {next_column} に書き込みました。")

def get_last_row(sheet, column_letter, search_min_row):
    """
    指定した列の最終行を取得する関数。
    
    sheet: 対象のシート
    column_letter: 対象の列（例：'A'、'B'など）
    """
    # 最後の行を取得
    for row in range(sheet.max_row, search_min_row-1, -1):  # 行を逆順にループ
        if sheet.cell(row=row, column=column_letter).value is not None:  # 値がある場合
            return row
    return 0  # 値が見つからない場合

def get_last_row_in_multiple_columns(sheet, columns):
    last_row = 1
    for column_letter in columns:
        last_row = max(
            last_row,
            get_last_row(
                sheet=sheet,
                column_letter=column_letter,
                search_min_row=last_row,
            ),
        )
    return last_row

def write_api_history_to_excel(
    date_time,
    value,
    model,
    excel_path,
    columns = [1, 2, 3]
):
    # 既存のExcelファイルを読み込む
    workbook = load_workbook(excel_path)
    sheet = workbook.active  # 操作対象のシートを指定 (デフォルトでアクティブシート)

    last_row = get_last_row_in_multiple_columns(sheet=sheet, columns=columns)
    next_row = last_row + 1
    
    values = [date_time, model, value]

    for i in range(3):
        sheet.cell(row=next_row, column=columns[i], value=values[i])

    # ファイルを保存
    workbook.save(excel_path)
